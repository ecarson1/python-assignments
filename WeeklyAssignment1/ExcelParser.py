from openpyxl import load_workbook
import datetime, logging, pathlib

# Helper to parse filename for desired month and year
def parse_file(file):
    args = file.split("_")
    month = args[3][:3] # first 3 letters of month
    month = datetime.datetime.strptime(month, "%b").month # gets int value of month
    year = int(args[4].split('.')[0]) # removes .xlsx leaving only year

    return (month, year)

# Helper to get column number for specific year and month in VOC sheet
def get_voc_column(ws, month, year):
    row = 1
    col = 2
    cell_value = ws.cell(row,col).value

    while cell_value != None:
        try:
            if cell_value.month == month and cell_value.year == year:
                return col

            col += 1
            cell_value = ws.cell(row,col).value
        except:
            col += 1
            cell_value = ws.cell(row,col).value

    raise ValueError('Month not found') 

# Helper to get row number for specific year and month in Summary sheet
def get_summary_row(ws, month, year):
    row = 2
    col = 1
    cell_value = ws.cell(row,col).value

    while cell_value != None:
        try:
            if cell_value.month == month and cell_value.year == year:
                return row

            row += 1
            cell_value = ws.cell(row,col).value
        except:
            row += 1
            cell_value = ws.cell(row,col).value

    raise ValueError('Month not found')

if __name__ == '__main__':
    #file = 'expedia_report_monthly_april_2016.xlsx'
    file = input('Enter the filename: \n')
    logpath = 'WeeklyAssignment1/Excel/parser_logfile.log'
    logging.basicConfig(filename=logpath, format='%(asctime)s: [%(levelname)s]: %(message)s', datefmt="%H:%M", level=logging.DEBUG)

    with open(logpath, 'r+') as f:
        f.truncate(0)
    logging.info('Logfile was cleared')

    try:
        (month, year) = parse_file(file)
    except:
        print("Incorrect file format. Example: expedia_report_monthly_march_2017.xlsx")
        logging.error(f'Given file, {file}, was improperly formatted')
        exit(0)
    logging.info(f'Filename {file} was parsed to get month {month} of {year}')

    try:
        wb = load_workbook('WeeklyAssignment1/Excel/' + file) # current directory was in main folder
    except:
        print('File failed to load. Make sure the path is correct')
        logging.error(f'Tried to load workbook from {file}, but it failed')
        exit(0)
    logging.info(f'Workbook loaded from {file}')

    # Getting Summary Rolling MoM info
    ws = wb['Summary Rolling MoM']
    logging.info(f'Summary Rolling MoM sheet loaded')
    try:
        row = get_summary_row(ws, month, year)
    except:
        print(f'Row not found in \'Summary Rolling MoM\' sheet for month {month} of {year}. Make sure months are formatted as a date')
        logging.error(f'Row for month {month} of {year} in Summary sheet of {file} was not found')
        exit(0)
            
    calls = ws.cell(row, 2).value
    abandon_pct = ws.cell(row, 3).value * 100
    fcr = ws.cell(row, 4).value * 100
    dsat = ws.cell(row, 5).value * 100
    csat = ws.cell(row, 6).value * 100

    logging.info(f"Results for month {month} of {year}: Calls Offered: {calls}, Abandon after 30s: {abandon_pct}%, FCR: {fcr}%, DSAT: {dsat}%, CSAT: {csat}%")

    # Getting VOC Rolling MoM info
    ws = wb['VOC Rolling MoM']
    logging.info(f'VOC Rolling MoM sheet loaded')
    try:
        col = get_voc_column(ws, month, year)
    except:
        print(f'Column not found in \'VOC Rolling MoM\' sheet for month {month} of {year}. Make sure months are formatted as a date')
        logging.error(f'Column for month {month} of {year} in VOC sheet of {file} was not found')
        exit(0)

    pro_score = ws.cell(4, col).value
    pass_score = ws.cell(6, col).value
    det_score = ws.cell(8, col).value

    pro_grade = "Good" if pro_score >= 200 else "Bad"
    pass_grade = "Good" if pass_score >= 100 else "Bad"
    det_grade = "Good" if det_score >= 100 else "Bad"

    logging.info(f"Results for month {month} of {year}: Promoters Score: {pro_score}, Promoters Grade: {pro_grade}, Passives Score: {pass_score}, Passives grade: {pass_grade}, Detractors Score: {det_score}, Detractors Grade: {det_grade}")
    logging.info("Program ended")
    wb.close()