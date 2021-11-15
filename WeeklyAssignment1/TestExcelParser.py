import unittest

from openpyxl.reader.excel import load_workbook
from ExcelParser import parse_file, get_summary_row, get_voc_column

class TestHelpers(unittest.TestCase):

    def test_parse_file(self):
        tup0 = parse_file("expedia_report_monthly_april_2016.xlsx")
        tup1 = parse_file("expedia_report_monthly_jan_2018.xlsx")

        self.assertEqual(tup0[0], 4)
        self.assertEqual(tup0[1], 2016)
        self.assertEqual(tup1[0], 1)
        self.assertEqual(tup1[1], 2018)

        # Testing exception handling
        thrown = False
        try:
            parse_file("test.xlsx")
        except:
            thrown = True
        self.assertTrue(thrown)

    def test_get_summary_row(self):
        wb = load_workbook('WeeklyAssignment1/Excel/expedia_report_monthly_january_2018.xlsx')
        ws = wb['Summary Rolling MoM']

        row1 = get_summary_row(ws, 3, 2017)
        row2 = get_summary_row(ws, 1, 2018)
        self.assertEqual(row1, 2)
        self.assertEqual(row2, 12)

        thrown = False
        try:
            get_summary_row(ws, 2, 2017)
        except ValueError:
            thrown = True
        self.assertTrue(thrown)
        wb.close()

    def test_get_voc_column(self):
        wb = load_workbook('WeeklyAssignment1/Excel/expedia_report_monthly_january_2018.xlsx')
        ws = wb['VOC Rolling MoM']

        col1 = get_voc_column(ws, 2, 2018)
        col2 = get_voc_column(ws, 4, 2016)
        self.assertEqual(col1, 2)
        self.assertEqual(col2, 24)
        
        thrown = False
        try:
            get_summary_row(ws, 4, 2015)
        except ValueError:
            thrown = True
        self.assertTrue(thrown)
        wb.close()


if __name__ == '__main__':
    unittest.main()